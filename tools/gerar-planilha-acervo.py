"""Gera planilha de consulta a partir da consolidação local, sem acessar Google."""
import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
data = json.loads((ROOT / 'dados/historico_orcamentos.json').read_text())
book = Workbook()
intro = book.active
intro.title = 'Leia primeiro'
intro.append(['Capital Fornecedores — histórico de orçamentos'])
intro.append(['Fonte', 'Pasta MELHOR IDEIA 2026; extrações e confirmação de contexto pelo usuário.'])
intro.append(['Natureza', 'Preços cotados. Não comprovam contratação, pagamento ou economia.'])
intro.append(['Unidades', 'Base Papéis não informa unidade: não comparar fornecedores sem validar embalagem.'])
intro.append(['Serviços', 'ADS: três componentes globais, sem preço unitário; não misturar com materiais.'])
intro.append(['Arredondamento', 'Litoral: unitário e total preservados como impressos; não reaplicar o desconto.'])
intro.append(['Linhas', len(data['registros'])])
intro.append(['Documentos', len(data['documentos'])])
intro.append(['Google', 'Esta planilha é uma consolidação local. Nada foi gravado na base do app.'])
intro.append(['Rastreabilidade', 'As colunas fonte_url e arquivo_codigo apontam para o PDF e a extração.'])

def table(name, rows):
    sheet = book.create_sheet(name)
    columns = list(rows[0])
    sheet.append(columns)
    for row in rows:
        values = []
        for column in columns:
            value = row.get(column)
            if isinstance(value, (dict, list)):
                value = json.dumps(value, ensure_ascii=False)
            elif column == 'data' and value:
                value = datetime.strptime(value, '%Y-%m-%d')
            values.append(value)
        sheet.append(values)
    sheet.freeze_panes = 'A2'
    sheet.auto_filter.ref = sheet.dimensions
    for index, column in enumerate(columns, 1):
        width = 22
        if column in ('descricao', 'arquivo_nome', 'evidencias', 'pendencias'):
            width = 65
        elif column in ('empreendimento', 'razao_social_fornecedor', 'fonte_url', 'arquivo_codigo'):
            width = 42
        sheet.column_dimensions[get_column_letter(index)].width = width
        for cells in sheet.iter_cols(min_col=index, max_col=index, min_row=2):
            for cell in cells:
                if column == 'data':
                    cell.number_format = 'dd/mm/yyyy'
                elif column in ('preco_unitario_cotado', 'valor_total_linha', 'total_declarado', 'total_somado', 'diferenca_total_menos_qtd_unitario'):
                    cell.number_format = '"R$" #,##0.00'
                elif column in ('codigo_fornecedor', 'numero_orcamento', 'numero', 'cnpj_fornecedor', 'cnpj_empresa', 'arquivo_id', 'registro_id'):
                    cell.number_format = '@'
                if column == 'fonte_url' and cell.value:
                    cell.hyperlink = cell.value
                    cell.style = 'Hyperlink'
    return sheet

table('Resumo por Mega', data['resumo'])
table('Historico', data['registros'])
table('Documentos', data['documentos'])
for sheet in book:
    sheet.sheet_view.showGridLines = False
    for cell in sheet[1]:
        cell.fill = PatternFill('solid', fgColor='151E49')
        cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(vertical='center', wrap_text=True)
    sheet.row_dimensions[1].height = 34
intro.column_dimensions['A'].width = 28
intro.column_dimensions['B'].width = 110
for row in intro.iter_rows():
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical='top')
destination = ROOT / 'dados/historico_orcamentos.xlsx'
book.save(destination)
print(destination)
