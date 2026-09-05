"""Gera a planilha de auditoria a partir do retrato JSON, sem acessar o Google."""
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
data = json.loads((ROOT / 'dados/auditoria_cobertura.json').read_text())
book = Workbook()
intro = book.active
intro.title = 'Leia primeiro'
intro.append(['Auditoria de cobertura — Megas', '05/09/2026'])
intro.append(['Finalidade', 'Lista do que existe, do que falta no histórico local e do que exige conciliação. Não é uma nova carga de preços.'])
intro.append(['Cobertura', '45 arquivos nas pastas principal, Engenharia e PLANO; mais 2 fontes referenciadas com acesso restrito.'])
intro.append(['Consolidado existente', '21 orçamentos / 274 linhas. Nenhuma alteração nesses registros nesta auditoria.'])
intro.append(['Cotações candidatas', '31: 26 nos mapas principais e 5 proponentes de Engenharia. Contagem provisória, sem contar revisões como novos contratos.'])
intro.append(['Contrato Canaveral', '28 preços fixados no anexo, sem inferir quantidade comprada ou pagamento.'])
intro.append(['Base Google', 'Já contém Wi-Fi e utilities: 2 importações, 4 equalizações, 12 registros de proposta e 78 linhas de preço (43 cotadas).'])
intro.append(['Revisões e duplicatas', 'Uma cotação pode aparecer em mapa, proposta, EAP, contrato e OC. Conferir os vínculos antes de inserir.'])
intro.append(['Valores', 'Totais anuais/globais e preços unitários não são comparáveis sem escopo, período, quantidade e unidade.'])
intro.append(['Fontes restritas', 'Originais de Wi-Fi e utilities responderam HTTP 401; foram conferidos somente seus registros na base compartilhada.'])
intro.append(['Proteção dos dados', 'Não inclui os dados pessoais de assinaturas nem cópia integral da base Google.'])

def table(name, rows):
    sheet = book.create_sheet(name)
    columns = list(dict.fromkeys(key for row in rows for key in row))
    sheet.append(columns)
    for row in rows:
        sheet.append([json.dumps(row.get(c), ensure_ascii=False) if isinstance(row.get(c), (list, dict)) else row.get(c) for c in columns])
    sheet.freeze_panes = 'A2'
    sheet.auto_filter.ref = sheet.dimensions
    for index, column in enumerate(columns, 1):
        sheet.column_dimensions[get_column_letter(index)].width = 65 if column in ('nome', 'achado', 'acao', 'observacao', 'descricao') else 28
        for cells in sheet.iter_cols(min_col=index, max_col=index, min_row=2):
            for cell in cells:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                if column in ('url', 'fonte_url') and cell.value:
                    cell.hyperlink = cell.value
                    cell.style = 'Hyperlink'
                elif column in ('valor_documento', 'preco_contratual'):
                    cell.number_format = '"R$" #,##0.00'
                elif column in ('cnpj', 'numero', 'codigo_anexo', 'id', 'contrato'):
                    cell.number_format = '@'
    return sheet

files = [r for r in data['arquivos'] if not r['tipo'].endswith('.folder')]
table('Arquivos fora do historico', [r for r in files if r['historico_local'] != 'presente'])
table('Cotacoes candidatas', data['cotacoes_candidatas'])
table('Precos contratuais Canaveral', data['precos_contratuais_canaveral'])
table('Todas as fontes', files)
table('Fontes com acesso pendente', data['fontes_referenciadas'])
table('Base Google', [{'campo': k, 'valor': v} for k, v in data['base_google'].items()])
for sheet in book:
    sheet.sheet_view.showGridLines = False
    for cell in sheet[1]:
        cell.fill = PatternFill('solid', fgColor='151E49')
        cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(vertical='center', wrap_text=True)
    sheet.row_dimensions[1].height = 36
intro.column_dimensions['A'].width = 30
intro.column_dimensions['B'].width = 110
for row in intro.iter_rows(min_row=2):
    intro.row_dimensions[row[0].row].height = 42
    for cell in row:
        cell.alignment = Alignment(vertical='top', wrap_text=True)
destination = ROOT / 'dados/auditoria_cobertura.xlsx'
book.save(destination)
print(destination)
